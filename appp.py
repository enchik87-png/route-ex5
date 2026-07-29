import io
import json
import math
import re
import time as time_module
import urllib.parse
from datetime import datetime, timedelta, time

import folium
import gspread
import openpyxl
import pandas as pd
import requests
import streamlit as st
from streamlit_folium import st_folium
from gspread_formatting import cellFormat, color, format_cell_range
from oauth2client.service_account import ServiceAccountCredentials

st.set_page_config(
    page_title="Kalyx Despatch Terminal", page_icon="🏍️", layout="centered"
)

# --- GOOGLE SHEET CONFIGURATION ---
SHEET_ID = "1YZgwm11scCWdiH5-9RxBVe3kEk9obfNtlH4frIYZGSg"
SHEET_EXPORT_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
SHEET_NAME = "Despatch"

# --- PRESET START & END LOCATIONS ---
PRESET_LOCATIONS = {
    "Kalyx Consultants Sdn Bhd (Office)": "🏢 Kalyx Consultants (Icon City, Bukit Mertajam)",
    "Machang Bubok (Home)": "🏠 Machang Bubok, Bukit Mertajam",
    "Taman Sri Serdang, Bertam (Home)": "🏠 Taman Sri Serdang, Bertam, Kepala Batas",
}

PRESET_COORDS = {
    "Kalyx Consultants Sdn Bhd (Office)": (5.343, 100.433),
    "Machang Bubok (Home)": (5.338, 100.508),
    "Taman Sri Serdang, Bertam (Home)": (5.518, 100.440)
}

# --- PENANG FULL REGION POSTCODE & SUB-AREA GPS ---
PENANG_POSTCODE_GPS = {
    # --- PENANG ISLAND (North to South Sweep: 10xxx & 11xxx) ---
    "10000": (5.416, 100.332),  # George Town
    "10150": (5.420, 100.315),  # Pulau Tikus
    "10250": (5.435, 100.300),  # Tanjung Tokong
    "10350": (5.455, 100.300),  # Tanjung Bungah
    "11200": (5.470, 100.270),  # Batu Ferringhi
    "11400": (5.405, 100.280),  # Air Itam / Farlim
    "11500": (5.385, 100.280),  # Paya Terubong
    "11600": (5.390, 100.310),  # Jelutong
    "11700": (5.355, 100.300),  # Gelugor / USM
    "11900": (5.320, 100.280),  # Bayan Lepas / Bayan Baru
    "11950": (5.290, 100.260),  # Teluk Kumbar
    "11000": (5.350, 100.240),  # Balik Pulau

    # --- MAINLAND SOUTH TO NORTH (13xxx & 14xxx) ---
    "14300": (5.170, 100.480),  # Nibong Tebal
    "14200": (5.220, 100.490),  # Sungai Bakap
    "14400": (5.200, 100.500),  # Valdor
    "14110": (5.250, 100.440),  # Batu Kawan
    "14120": (5.280, 100.480),  # Simpang Ampat
    "14100_serijuru": (5.290, 100.430),
    "14100_tambun": (5.300, 100.440),
    "14100_bukitminyak": (5.310, 100.450),
    "14100": (5.305, 100.445),
    "14000_juru": (5.330, 100.435),
    "14000_teguh": (5.340, 100.450),
    "14000_bm": (5.350, 100.460),
    "14000": (5.350, 100.460),
    "13500": (5.358, 100.410),  # Permatang Pauh / Pauh Jaya
    "13600": (5.365, 100.390),  # Perai
    "13700": (5.390, 100.400),  # Seberang Jaya
    "13000": (5.412, 100.370),  # Butterworth
    "13020": (5.418, 100.380),  # Butterworth / Selayang Indah
    "13400": (5.420, 100.380),  # Butterworth / Bagan
    "13800": (5.445, 100.430),  # Sungai Dua
    "13300": (5.480, 100.480),  # Tasek Gelugor
    "13200": (5.515, 100.430),  # Kepala Batas
}


# --- 1. SESSION STATE MANAGEMENT ---
if "page" not in st.session_state:
    st.session_state.page = "setup"
if "optimized_route" not in st.session_state:
    st.session_state.optimized_route = []
if "current_stop" not in st.session_state:
    st.session_state.current_stop = 0
if "optimization_time" not in st.session_state:
    st.session_state.optimization_time = None
if "start_point" not in st.session_state:
    st.session_state.start_point = "Kalyx Consultants Sdn Bhd (Office)"
if "end_point" not in st.session_state:
    st.session_state.end_point = "Kalyx Consultants Sdn Bhd (Office)"


# --- 2. KPI AGING CALCULATION LOGIC ---
def calculate_kpi_status(raw_request_date):
    now = datetime.now()
    req_dt = None

    if isinstance(raw_request_date, datetime):
        req_dt = raw_request_date
    elif isinstance(raw_request_date, str) and raw_request_date.strip():
        try:
            req_dt = pd.to_datetime(raw_request_date).to_pydatetime()
        except Exception:
            req_dt = now

    if not req_dt:
        req_dt = now

    if req_dt.time() > time(9, 30):
        effective_date = req_dt.date() + timedelta(days=1)
    else:
        effective_date = req_dt.date()

    days_elapsed = (now.date() - effective_date).days
    days_elapsed = max(0, days_elapsed)

    if days_elapsed >= 3:
        status_key = "Overdue"
        badge_html = (
            f"<span style='background-color:#FF4D4D; color:white; padding:3px 8px; "
            f"border-radius:4px; font-weight:bold; font-size:12px;'>"
            f"🚨 OVERDUE ({days_elapsed} Days)</span>"
        )
        is_red = True
    elif days_elapsed == 2:
        status_key = "Due Today (Day 2)"
        badge_html = (
            "<span style='background-color:#FFC107; color:black; padding:3px 8px; "
            "border-radius:4px; font-weight:bold; font-size:12px;'>"
            "⚠️ KPI LIMIT (Day 2)</span>"
        )
        is_red = False
    else:
        day_label = "Day" if days_elapsed == 1 else "Days"
        status_key = "On Time"
        badge_html = (
            f"<span style='background-color:#28A745; color:white; padding:3px 8px; "
            f"border-radius:4px; font-weight:bold; font-size:12px;'>"
            f"✅ ON TIME ({days_elapsed} {day_label})</span>"
        )
        is_red = False

    formatted_req_str = req_dt.strftime("%d/%m/%Y %I:%M %p")
    return days_elapsed, status_key, badge_html, is_red, formatted_req_str


def _normalize_area_name(raw_area):
    cleaned = str(raw_area).strip().title()
    if not cleaned or cleaned.lower() in {"nan", "none", "-"}:
        return ""
    return cleaned

def get_cluster_key(raw_area, address):
    combined_text = f"{str(raw_area)} {str(address)}"
    postcode_match = re.search(r"\b(\d{5})\b", combined_text)
    if postcode_match:
        pc = postcode_match.group(1)
        region = "Island" if pc.startswith(("10", "11")) else "Mainland"
        return f"[{region}] Postcode {pc}"
    cleaned = _normalize_area_name(raw_area)
    if cleaned:
        return f"Area: {cleaned}"
    return "Unassigned"

def _safe_float(value, default=0.0):
    try:
        if value is None or str(value).strip().lower() in {"", "nan", "-", "tbc"}:
            return default
        return float(value)
    except (TypeError, ValueError):
        return default


# --- 3. DATA LOADING LOGIC ---
@st.cache_data(ttl=30, show_spinner=False)
def load_pending_despatch_tasks():
    response = requests.get(SHEET_EXPORT_URL, timeout=20)
    if response.status_code != 200:
        raise Exception("Failed to download Google Sheet.")

    wb = openpyxl.load_workbook(io.BytesIO(response.content), data_only=True)
    sheet = wb[SHEET_NAME]
    pending_tasks = []
    two_weeks_ago = datetime.now() - timedelta(days=14)

    for row_idx in range(2, sheet.max_row + 1):
        cell_date = sheet.cell(row=row_idx, column=1).value
        task_date = None
        if isinstance(cell_date, datetime):
            task_date = cell_date
        elif isinstance(cell_date, str):
            try:
                task_date = pd.to_datetime(cell_date)
            except Exception:
                pass

        if task_date:
            if hasattr(task_date, "to_pydatetime"):
                task_date = task_date.to_pydatetime()
            if task_date < two_weeks_ago:
                continue

        cell_address = sheet.cell(row=row_idx, column=19)
        fill = cell_address.fill
        is_orange = False
        if fill and fill.start_color and fill.start_color.rgb:
            cell_color = str(fill.start_color.rgb).upper()
            if any(c in cell_color for c in ["A500", "9900", "FFC000", "ED7D31", "F290"]):
                is_orange = True

        if not is_orange:
            address = cell_address.value
            company = sheet.cell(row=row_idx, column=17).value

            if company:
                days_elapsed, kpi_status, badge_html, is_red, formatted_req_str = calculate_kpi_status(cell_date)
                raw_area = str(sheet.cell(row=row_idx, column=14).value or "Unassigned")
                cluster_group = get_cluster_key(raw_area, str(address))

                pending_tasks.append(
                    {
                        "id": row_idx,
                        "requested_date_raw": cell_date,
                        "requested_date_str": formatted_req_str,
                        "kpi_days": days_elapsed,
                        "kpi_status": kpi_status,
                        "kpi_badge_html": badge_html,
                        "is_red_overdue": is_red,
                        "pic_name": str(sheet.cell(row=row_idx, column=5).value or "-"),
                        "task_type": str(sheet.cell(row=row_idx, column=6).value or "Despatch"),
                        "area": cluster_group,
                        "box": _safe_float(sheet.cell(row=row_idx, column=16).value),
                        "company": str(company),
                        "address": str(address),
                        "target_date": str(sheet.cell(row=row_idx, column=23).value or "-")[:10],
                        "container": _safe_float(sheet.cell(row=row_idx, column=24).value),
                        "bag": _safe_float(sheet.cell(row=row_idx, column=25).value),
                        "envelope": _safe_float(sheet.cell(row=row_idx, column=26).value),
                        "transport": str(sheet.cell(row=row_idx, column=27).value or "Motorcycle"),
                        "sheet_time_slot": str(sheet.cell(row=row_idx, column=28).value or "").strip(),
                        "department": str(sheet.cell(row=row_idx, column=37).value or "-"),
                        "client": str(sheet.cell(row=row_idx, column=38).value or "N/A"),
                        "phone": str(sheet.cell(row=row_idx, column=39).value or ""),
                    }
                )
    return pending_tasks


# --- 4. GOOGLE SHEETS WRITEBACK ---
def mark_row_completed_in_sheets(row_idx):
    if "gcp_service_account" not in st.secrets:
        return "no_secrets"
    try:
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(
            dict(st.secrets["gcp_service_account"]),
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive",
            ],
        )
        gc = gspread.authorize(credentials)
        sh = gc.open_by_key(SHEET_ID)
        worksheet = sh.worksheet(SHEET_NAME)
        fmt = cellFormat(backgroundColor=color(1.0, 0.65, 0.0))
        format_cell_range(worksheet, f"A{row_idx}:AM{row_idx}", fmt)
        return True
    except Exception:
        return False


# --- 5. ROUTING & MAP RENDERING ---
def get_stop_coords(stop):
    combined_text = f"{stop.get('area', '')} {stop.get('address', '')}".lower()
    match = re.search(r"\b(1\d{4})\b", combined_text)
    
    if match:
        pc = match.group(1)
        if pc == "14000":
            if "juru" in combined_text:
                return PENANG_POSTCODE_GPS["14000_juru"]
            elif any(k in combined_text for k in ["teguh", "permatang tinggi"]):
                return PENANG_POSTCODE_GPS["14000_teguh"]
            else:
                return PENANG_POSTCODE_GPS["14000_bm"]
        elif pc == "14100":
            if "seri juru" in combined_text:
                return PENANG_POSTCODE_GPS["14100_serijuru"]
            elif "tambun" in combined_text:
                return PENANG_POSTCODE_GPS["14100_tambun"]
            else:
                return PENANG_POSTCODE_GPS["14100_bukitminyak"]
        elif pc in PENANG_POSTCODE_GPS:
            return PENANG_POSTCODE_GPS[pc]

    return (5.350, 100.350)

def optimize_route_with_gemini(stops_list, start_key, end_key):
    if not stops_list:
        return []
    sorted_stops = sorted(
        stops_list, 
        key=lambda s: (get_stop_coords(s)[0], s.get("company", ""))
    )
    return sorted_stops

def render_interactive_map(route_list, start_key, end_key):
    start_coords = PRESET_COORDS.get(start_key, (5.343, 100.433))
    
    # Center map around start point or first stop
    m = folium.Map(location=start_coords, zoom_start=12, tiles="OpenStreetMap")

    # Start Point Marker
    folium.Marker(
        location=start_coords,
        popup=f"<b>START:</b> {start_key}",
        icon=folium.Icon(color="green", icon="play", prefix="fa")
    ).add_to(m)

    coord_points = [start_coords]

    for i, stop in enumerate(route_list):
        coords = get_stop_coords(stop)
        coord_points.append(coords)
        
        popup_html = f"<b>Stop #{i+1}: {stop['company']}</b><br>{stop['address']}"
        folium.Marker(
            location=coords,
            popup=popup_html,
            tooltip=f"#{i+1}: {stop['company']}",
            icon=folium.DivIcon(html=f"""
                <div style="background-color: #1E90FF; color: white; border-radius: 50%; 
                            width: 24px; height: 24px; display: flex; align-items: center; 
                            justify-content: center; font-weight: bold; font-size: 11px; 
                            border: 2px solid white; box-shadow: 0px 2px 4px rgba(0,0,0,0.3);">
                    {i+1}
                </div>
            """)
        ).add_to(m)

    # Draw connecting polyline path
    if len(coord_points) > 1:
        folium.PolyLine(coord_points, color="#1E90FF", weight=4, opacity=0.8).add_to(m)

    st_folium(m, height=350, use_container_width=True)


# --- 6. UI PAGE: SETUP & TASK SELECTION ---
if st.session_state.page == "setup":
    st.title("🏍️ Kalyx Despatch Terminal")

    with st.spinner("Fetching live pending tasks..."):
        try:
            tasks = load_pending_despatch_tasks()
        except Exception as e:
            st.error(f"Error loading sheet: {e}")
            st.stop()

    if not tasks:
        st.success("🎉 All recent tasks are completed!")
        st.stop()

    df_tasks = pd.DataFrame(tasks)

    col_f1, col_f2, col_f3 = st.columns(3)
    with col_f1:
        selected_area = st.selectbox("📍 Filter Zone/Postcode:", ["All Zones"] + list(df_tasks["area"].unique()))
    with col_f2:
        selected_transport = st.selectbox("🚗/🏍️ Transport:", ["All", "Car", "Motorcycle"])
    with col_f3:
        selected_kpi = st.selectbox(
            "⏳ KPI Filter:",
            ["All Tasks", "🚨 Overdue Only", "⚠️ Due Today (Day 2)", "✅ On Time Only"],
        )

    filtered_df = df_tasks
    if selected_area != "All Zones":
        filtered_df = filtered_df[filtered_df["area"] == selected_area]
    if selected_transport != "All":
        filtered_df = filtered_df[filtered_df["transport"].str.contains(selected_transport, case=False)]
    if selected_kpi == "🚨 Overdue Only":
        filtered_df = filtered_df[filtered_df["is_red_overdue"]]
    elif selected_kpi == "⚠️ Due Today (Day 2)":
        filtered_df = filtered_df[filtered_df["kpi_status"] == "Due Today (Day 2)"]
    elif selected_kpi == "✅ On Time Only":
        filtered_df = filtered_df[filtered_df["kpi_status"] == "On Time"]

    st.markdown(f"### 📋 Select Tasks for Today ({len(filtered_df)} Available)")

    selected_stops = []

    for _, row in filtered_df.iterrows():
        row_dict = row.to_dict()

        items = []
        if row["box"] > 0:
            items.append(f"📦 {int(row['box'])} Box")
        if row["container"] > 0:
            items.append(f"🗃️ {int(row['container'])} Container")
        if row["bag"] > 0:
            items.append(f"🛍️ {int(row['bag'])} Bag")
        if row["envelope"] > 0:
            items.append(f"✉️ {int(row['envelope'])} Envelope")
        items_str = " | ".join(items) if items else "No document details"

        border_style = "border-left: 5px solid #FF4D4D; padding-left: 8px;" if row["is_red_overdue"] else ""

        c_check, c_details, c_time = st.columns([0.08, 0.64, 0.28])

        with c_check:
            is_checked = st.checkbox("Select", key=f"chk_{row['id']}")

        with c_details:
            transport_icon = "🚗" if "car" in str(row["transport"]).lower() else "🏍️"
            st.markdown(
                f"<div style='{border_style}'>"
                f"<b>{row['company']}</b> {transport_icon} - <i>{row['task_type']}</i> {row['kpi_badge_html']}<br>"
                f"<small>📅 <b>Requested:</b> {row['requested_date_str']}</small><br>"
                f"<small>🏢 <b>Dept:</b> {row['department']} | 👤 <b>PIC:</b> {row['pic_name']} | 🤝 <b>Client:</b> {row['client']}</small><br>"
                f"<small>📄 <b>Items:</b> {items_str}</small><br>"
                f"<small>📍 [{row['area']}] {row['address']}</small>"
                f"</div>",
                unsafe_allow_html=True,
            )

        with c_time:
            has_time = st.checkbox("Set Time", key=f"time_chk_{row['id']}")
            job_time = None
            if has_time:
                job_time = st.time_input(
                    "Target Time",
                    value=time(9, 0),
                    key=f"t_val_{row['id']}",
                    label_visibility="collapsed",
                )
            row_dict["custom_time"] = job_time

        if is_checked:
            selected_stops.append(row_dict)
        st.markdown("---")

    st.markdown("### 🗺️ Route Endpoints & Optimization")
    col_sp, col_ep = st.columns(2)
    options_keys = list(PRESET_LOCATIONS.keys())

    with col_sp:
        sel_start = st.selectbox("🚩 Start Point:", options_keys, index=0, key="select_start_pt")
    with col_ep:
        sel_end = st.selectbox("🏁 End Point:", options_keys, index=0, key="select_end_pt")

    if st.button("⚡ Calculate Smart Route", type="primary", use_container_width=True):
        if not selected_stops:
            st.warning("Please select at least one task.")
        else:
            st.session_state.optimization_time = datetime.now()
            st.session_state.start_point = sel_start
            st.session_state.end_point = sel_end

            with st.spinner("⚡ Processing Penang routing & map generation..."):
                st.session_state.optimized_route = optimize_route_with_gemini(selected_stops, sel_start, sel_end)

            st.session_state.page = "preview"
            st.rerun()


# --- 7. UI PAGE: ROUTE PREVIEW & MAP ---
elif st.session_state.page == "preview":
    st.title("🗺️ Route Preview & Map")
    st.caption("Review your optimized route sequence and visual map below.")

    start_label = PRESET_LOCATIONS.get(st.session_state.start_point, st.session_state.start_point)
    end_label = PRESET_LOCATIONS.get(st.session_state.end_point, st.session_state.end_point)
    st.info(f"🚩 **Start:** {start_label}  \n🏁 **End:** {end_label}")

    # Render In-App Map
    st.markdown("#### 🗺️ Route Overview Map")
    render_interactive_map(st.session_state.optimized_route, st.session_state.start_point, st.session_state.end_point)
    st.markdown("---")

    route_list = st.session_state.optimized_route
    total_stops = len(route_list)

    for i, stop in enumerate(route_list):
        transport_icon = "🚗" if "car" in str(stop["transport"]).lower() else "🏍️"
        c_move, c_info = st.columns([0.22, 0.78])

        with c_move:
            st.markdown(f"**Stop #{i + 1}**")
            col_u, col_d = st.columns(2)
            with col_u:
                if i > 0 and st.button("⬆️", key=f"up_{i}"):
                    route_list[i], route_list[i - 1] = route_list[i - 1], route_list[i]
                    st.rerun()
            with col_d:
                if i < total_stops - 1 and st.button("⬇️", key=f"down_{i}"):
                    route_list[i], route_list[i + 1] = route_list[i + 1], route_list[i]
                    st.rerun()

        with c_info:
            st.markdown(
                f"<b>{stop['company']}</b> {transport_icon} {stop['kpi_badge_html']}<br>"
                f"<small>📍 <b>[{stop['area']}]</b> {stop['address']}</small>",
                unsafe_allow_html=True,
            )
        st.markdown("---")

    col_btn1, col_btn2 = st.columns(2)
    with col_btn1:
        if st.button("⬅️ Re-select Tasks", use_container_width=True):
            st.session_state.page = "setup"
            st.rerun()
    with col_btn2:
        if st.button("🚀 Start Live Route", type="primary", use_container_width=True):
            st.session_state.current_stop = 0
            st.session_state.page = "route"
            st.rerun()


# --- 8. UI PAGE: LIVE ROUTE EXECUTION ---
elif st.session_state.page == "route":
    total_stops = len(st.session_state.optimized_route)
    current_idx = st.session_state.current_stop
    stop = st.session_state.optimized_route[current_idx]

    col_nav_top, col_home = st.columns([0.75, 0.25])
    with col_nav_top:
        opt_time_str = (
            st.session_state.optimization_time.strftime("%I:%M %p")
            if st.session_state.optimization_time
            else ""
        )
        st.markdown(f"### 🏁 Stop {current_idx + 1} of {total_stops}")
        st.caption(f"Optimized at {opt_time_str}")
    with col_home:
        if st.button("🏠 Home", key="btn_home_route", use_container_width=True):
            st.session_state.page = "setup"
            st.session_state.current_stop = 0
            st.session_state.optimized_route = []
            st.rerun()

    st.progress((current_idx + 1) / total_stops)
    st.markdown("---")

    transport_icon = "🚗" if "car" in str(stop["transport"]).lower() else "🏍️"

    st.title(f"{stop['company']}")
    st.markdown(f"{stop['kpi_badge_html']}", unsafe_allow_html=True)
    time_badge = (
        f"⏰ Slot: {stop['custom_time'].strftime('%I:%M %p')}"
        if stop.get("custom_time")
        else "⏰ Flexible / Anytime"
    )
    st.markdown(f"**Task:** {stop['task_type']} {transport_icon} | {time_badge}")
    st.markdown(f"📅 **Requested Date:** {stop['requested_date_str']}")
    st.markdown(f"🏢 **Dept:** {stop['department']} | 👤 **Kalyx PIC:** {stop['pic_name']}")
    st.markdown(f"📍 **Zone/Postcode:** {stop['area']} | **Address:** {stop['address']}")

    items = []
    if stop["box"] > 0:
        items.append(f"📦 {int(stop['box'])} Box")
    if stop["container"] > 0:
        items.append(f"🗃️ {int(stop['container'])} Container")
    if stop["bag"] > 0:
        items.append(f"🛍️ {int(stop['bag'])} Bag")
    if stop["envelope"] > 0:
        items.append(f"✉️ {int(stop['envelope'])} Envelope")
    if items:
        st.info(" | ".join(items))

    clean_phone = str(stop["phone"]).replace(" ", "").replace("-", "") if stop["phone"] else ""
    col_c1, col_c2 = st.columns(2)

    with col_c1:
        if clean_phone and clean_phone != "nan":
            st.markdown(
                f"<a href='tel:{clean_phone}'><button style='width: 100%; padding:14px; border-radius:8px; border:1px solid #1E90FF; background:transparent; color:#1E90FF; font-weight:bold;'>📞 Call Client ({stop['client']})</button></a>",
                unsafe_allow_html=True,
            )
        else:
            st.write(f"📞 Client Contact: {stop['client']} (No phone)")

    with col_c2:
        search_query = f"{stop['company']}, {stop['address']}, Penang, Malaysia"
        maps_url = f"https://www.google.com/maps/dir/?api=1&destination={urllib.parse.quote(search_query)}"
        st.markdown(
            f"<a href='{maps_url}' target='_blank'><button style='width: 100%; padding:14px; border-radius:8px; border:1px solid #28a745; background:transparent; color:#28a745; font-weight:bold;'>🗺️ Navigate</button></a>",
            unsafe_allow_html=True,
        )

    st.write("")
    if st.button("✅ Mark Complete & Show Next Job", type="primary", use_container_width=True):
        with st.spinner("Updating Google Sheets..."):
            success = mark_row_completed_in_sheets(stop["id"])

            if success == "no_secrets":
                st.warning("⚠️ Google Sheets API not connected yet. Moving to next job in the app only.")
                st.session_state.current_stop += 1
                if st.session_state.current_stop >= total_stops:
                    st.session_state.page = "finished"
                st.rerun()

            elif success is True:
                st.session_state.current_stop += 1
                if st.session_state.current_stop >= total_stops:
                    st.session_state.page = "finished"
                st.rerun()

            else:
                st.error("❌ Failed to update Google Sheets. Check your internet connection and try again.")


# --- 9. UI PAGE: SHIFT COMPLETED ---
elif st.session_state.page == "finished":
    st.balloons()
    st.title("🎉 All Tasks Completed!")
    st.success("All selected jobs for this route have been completed and updated live in Google Sheets.")

    col_f1, col_f2 = st.columns(2)
    with col_f1:
        if st.button("🔄 Start New Shift", use_container_width=True):
            st.session_state.page = "setup"
            st.cache_data.clear()
            st.rerun()
    with col_f2:
        if st.button("🏠 Home", key="btn_home_finished", use_container_width=True):
            st.session_state.page = "setup"
            st.cache_data.clear()
            st.rerun()
